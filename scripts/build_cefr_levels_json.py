#!/usr/bin/env python3
"""
Rebuild web/src/data/cefr-a1-a2.json from CEFR xlsx exports (A1, A2, optional B1).

Usage:
  python3 scripts/build_cefr_levels_json.py <A1.xlsx> <A2.xlsx> [B1.xlsx]

Sheets: A1/A2 use sheet "Vocabulary" column A (legacy). B1 uses "word_index"
or first sheet, column "headword" (column A).

Requires: pip install openpyxl
"""
from __future__ import annotations

import json
import sys
from pathlib import Path


def load_a1_a2_sheet(path: Path) -> set[str]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Vocabulary"]
    s: set[str] = set()
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        w = row[0]
        if w and str(w).strip():
            s.add(str(w).strip().lower())
    wb.close()
    return s


def load_b1_or_flexible(path: Path) -> set[str]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "word_index" in wb.sheetnames:
        ws = wb["word_index"]
    elif "Vocabulary" in wb.sheetnames:
        ws = wb["Vocabulary"]
    else:
        ws = wb[wb.sheetnames[0]]
    s: set[str] = set()
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        w = row[0]
        if w is None:
            continue
        t = str(w).strip().lower()
        if t:
            s.add(t)
    wb.close()
    return s


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    out = root / "web" / "src" / "data" / "cefr-a1-a2.json"
    if len(sys.argv) < 3:
        print(
            "Usage: python3 build_cefr_levels_json.py <A1.xlsx> <A2.xlsx> [B1.xlsx]",
            file=sys.stderr,
        )
        sys.exit(1)
    a1p = Path(sys.argv[1]).resolve()
    a2p = Path(sys.argv[2]).resolve()
    if not a1p.is_file() or not a2p.is_file():
        print("A1 or A2 xlsx not found.", file=sys.stderr)
        sys.exit(1)
    a1 = load_a1_a2_sheet(a1p)
    a2 = load_a1_a2_sheet(a2p)
    files = [a1p.name, a2p.name]
    b1: set[str] = set()
    if len(sys.argv) >= 4:
        b1p = Path(sys.argv[3]).resolve()
        if not b1p.is_file():
            print("B1 xlsx not found.", file=sys.stderr)
            sys.exit(1)
        b1 = load_b1_or_flexible(b1p)
        files.append(b1p.name)
    data: dict = {
        "version": 2 if b1 else 1,
        "sourceFiles": files,
        "note": "Lookup: trim+lower. Priority: A1, then A2, then B1 (first hit wins).",
        "a1Count": len(a1),
        "a2Count": len(a2),
        "a1": sorted(a1),
        "a2": sorted(a2),
    }
    if b1:
        data["b1Count"] = len(b1)
        data["overlapA1B1"] = len(a1 & b1)
        data["overlapA2B1"] = len(a2 & b1)
        data["b1"] = sorted(b1)
    else:
        data["overlapCount"] = len(a1 & a2)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=0)
    print("Wrote", out, f"({out.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
